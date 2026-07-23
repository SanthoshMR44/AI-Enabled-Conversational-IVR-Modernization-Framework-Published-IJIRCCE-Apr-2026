import os
import json
import csv
from datetime import datetime
from flask import Blueprint, jsonify, request, send_file, current_app
from database import get_db

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/api/analytics/dashboard', methods=['GET'])
def dashboard_stats():
    conn = get_db()
    
    total_calls = conn.execute("SELECT COUNT(*) FROM conversations").fetchone()[0]
    avg_duration = conn.execute("SELECT AVG(duration) FROM conversations").fetchone()[0] or 0
    avg_sentiment = conn.execute("SELECT AVG(sentiment_score) FROM conversations").fetchone()[0] or 0.0
    
    # Sentiment count
    pos_calls = conn.execute("SELECT COUNT(*) FROM conversations WHERE sentiment = 'Positive'").fetchone()[0]
    neu_calls = conn.execute("SELECT COUNT(*) FROM conversations WHERE sentiment = 'Neutral'").fetchone()[0]
    neg_calls = conn.execute("SELECT COUNT(*) FROM conversations WHERE sentiment = 'Negative'").fetchone()[0]
    
    # Intent distribution
    intent_rows = conn.execute("SELECT intent_detected, COUNT(*) as count FROM conversations GROUP BY intent_detected").fetchall()
    intents_dist = {r['intent_detected']: r['count'] for r in intent_rows if r['intent_detected']}
    
    # Recent logs
    logs_rows = conn.execute("SELECT * FROM conversations ORDER BY start_time DESC LIMIT 10").fetchall()
    logs = [dict(r) for r in logs_rows]
    
    # Accuracy estimation metrics
    correct_classifications = conn.execute("SELECT COUNT(*) FROM conversations WHERE intent_detected != 'Unknown Intent'").fetchone()[0]
    total_non_empty = max(total_calls, 1)
    accuracy = round((correct_classifications / total_non_empty) * 100, 2)
    
    conn.close()
    
    return jsonify({
        "total_calls": total_calls,
        "avg_duration": round(avg_duration, 1),
        "avg_sentiment": round(avg_sentiment, 2),
        "sentiment_distribution": {
            "Positive": pos_calls,
            "Neutral": neu_calls,
            "Negative": neg_calls
        },
        "intent_distribution": intents_dist,
        "recent_logs": logs,
        "accuracy": accuracy
    })

@reports_bp.route('/api/reports/export', methods=['GET'])
def export_reports():
    file_format = request.args.get('format', 'csv')
    conn = get_db()
    rows = conn.execute("SELECT * FROM conversations ORDER BY start_time DESC").fetchall()
    conn.close()
    
    # Temporary file output
    temp_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'scratch')
    os.makedirs(temp_dir, exist_ok=True)
    
    if file_format == 'csv':
        filepath = os.path.join(temp_dir, "conversation_report.csv")
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Call ID", "Caller Number", "Start Time", "End Time", "Duration (s)", "Sentiment", "Sentiment Score", "Intent Detected", "Entities"])
            for r in rows:
                writer.writerow([r['call_id'], r['caller_number'], r['start_time'], r['end_time'], r['duration'], r['sentiment'], r['sentiment_score'], r['intent_detected'], r['entities']])
        return send_file(filepath, as_attachment=True, download_name="conversation_report.csv")
        
    elif file_format == 'excel':
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Conversations"
        ws.append(["Call ID", "Caller Number", "Start Time", "End Time", "Duration (s)", "Sentiment", "Sentiment Score", "Intent Detected", "Entities"])
        for r in rows:
            ws.append([r['call_id'], r['caller_number'], r['start_time'], r['end_time'], r['duration'], r['sentiment'], r['sentiment_score'], r['intent_detected'], r['entities']])
        filepath = os.path.join(temp_dir, "conversation_report.xlsx")
        wb.save(filepath)
        return send_file(filepath, as_attachment=True, download_name="conversation_report.xlsx")
        
    else:  # PDF Export
        from fpdf import FPDF
        class PDF(FPDF):
            def header(self):
                self.set_font('Arial', 'B', 15)
                self.cell(80)
                self.cell(30, 10, 'Conversational IVR Platform Report', 0, 0, 'C')
                self.ln(20)
            def footer(self):
                self.set_y(-15)
                self.set_font('Arial', 'I', 8)
                self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')
                
        pdf = PDF()
        pdf.alias_nb_pages()
        pdf.add_page()
        pdf.set_font('Times', '', 12)
        pdf.cell(0, 10, f"Generated At: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 0, 1)
        pdf.cell(0, 10, f"Total Calls Analyzed: {len(rows)}", 0, 1)
        pdf.ln(10)
        
        for i, r in enumerate(rows[:50]): # Cap at 50 for layout simplicity
            pdf.set_font('Times', 'B', 10)
            pdf.cell(0, 6, f"Call ID: {r['call_id']} | Caller: {r['caller_number']} | Duration: {r['duration']}s", 0, 1)
            pdf.set_font('Times', '', 9)
            pdf.cell(0, 5, f"Intent: {r['intent_detected']} | Sentiment: {r['sentiment']} (Score: {r['sentiment_score']})", 0, 1)
            pdf.multi_cell(0, 5, f"Transcript: {r['transcript']}")
            pdf.ln(3)
            
        filepath = os.path.join(temp_dir, "conversation_report.pdf")
        pdf.output(filepath, 'F')
        return send_file(filepath, as_attachment=True, download_name="conversation_report.pdf")
